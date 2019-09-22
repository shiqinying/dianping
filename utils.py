import openpyxl
import requests
from openpyxl import Workbook
from config import *


# 从excle表获取所有店名
def get_shop_name():
    # 打开excel文件,获取工作簿对象
    wb = openpyxl.load_workbook('店名20000.xlsx')
    # 从表单中获取单元格的内容
    ws = wb.active  # 当前活跃的表单
    list = []
    for i in range(2,22603):
        value = ws.cell(row=i, column=2).value
        list.append(value.strip())
    print(list)

# 根据店名获取经纬度
def name2location(name):
    url = f'http://api.map.baidu.com/geocoding/v3/?address={name}&output=json&ak={BAIDU_AK}&city=beijing'
    r = requests.get(url=url).json()
    print(r)
    return r

def get_geo():
    mongo = Mongo('shop_20000')
    for name in SHOP_NAME:

        r = name2location(name.strip())
        if r.get('status') == 0:
            result = r.get('result')
            item = {}
            item['shop_name'] = name
            item['lng'] = result['location']['lng']
            item['lat'] = result['location']['lat']
            item['precise'] = result['precise']
            item['confidence'] = result['confidence']
            item['level'] = result['level']
            print(item)
            mongo.insert('shop_geo',item,'shop_name')
        else:
            continue

def save2xlsx(ws):
    mongo = Mongo('shop_20000')
    results = mongo.findall('shop_geo')
    row = 1
    for i in results:
        item = {}
        item['shop_name'] = i.get('shop_name', '').strip()
        item['lng'] = str(i.get('lng', '')).strip()
        item['lat'] = str(i.get('lat', '')).strip()
        item['precise'] = str(i.get('precise', '')).strip()
        item['confidence'] = str(i.get('confidence', '')).strip()
        item['level'] = str(i.get('level', '')).strip()
        row += 1
        c_list = ['shop_name','lng','lat','precise','confidence','level']
        for c in range(0, len(c_list)):
            key = c_list[c]
            value = item[key]
            ws.cell(row=row, column=c+1, value=value)

def run(func,to_file, title='sheet0'):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    func(ws)
    wb.save(filename=to_file)


if __name__=='__main__':
    # get_shop_name()
    # get_geo()
    # name = '星期六'
    # name2location(name)
    # run(func=save2xlsx, to_file='shop_geo.xlsx')

    mongo = Mongo('shop_20000')
    shops = mongo.findall('dianping_shop_id')
    for shop in shops:
        REDIS.sadd('shop_id',shop['shop_id'])